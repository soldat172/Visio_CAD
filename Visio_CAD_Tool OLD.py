import pyautogui, time, openpyxl, keyboard
from openpyxl import Workbook
from tkinter import filedialog, Tk
from guizero import App, Text, TextBox, PushButton, Window

#Declaring of variables
global siteName
siteName = str ('Paul')
global apPrefix
apPrefix = siteName + '-XX-'
global apNumber
apNumber =  int('1')
global apName
apName = str(apPrefix) + str(apNumber)
global count
count = int('1')
global excelNumber
excelNumber = int('1')
global name
name = 'A'
global number
number = int('1')
global rowNumber
rowNumber = str(name) + str(number)
global list
list=[99999]
global excelName
excelName = 'Paul'
global formatNumber
formatNumber = format(apNumber, '05')
global excelStartingInteger
excelStartingInteger = int('1')

def startWorkbook (): #Starts an excel. Required for visioTool.
    global wb
    wb = Workbook()
    global ws
    ws = wb.active

def openFiles(): #Opens file explorer
    global excelName
    Tk().withdraw() # we don't want a full GUI, so keep the root window from appearing
    excelName = filedialog.askopenfilename() # show an "Open" dialog box and return the path to the selected file
    print(excelName)
    print('')

def visioTool():
    global apName
    global apNumber
    global formatNumber
    global excelStartingInteger
#----------------------------------------------------------------------Makes GUI work
    global app
    global window2
    global displayText
    displayText = Text(window2, text=apName, align="left", grid=[0,0])
#----------------------------------------------------------------------Makes GUI work
    def visioGuts (): #Controls the actions of the Visio Tool
        global apNumber
        global formatNumber
        global excelStartingInteger
#----------------------------------------------------------------------Makes GUI work
        global app
        global window2
        global displayText
        displayText.clear()
        window2.focus()
        displayText.append(apName)
        window3 = Window(app, visible=False)
#----------------------------------------------------------------------Makes GUI work
        print (apName)
        ws.cell(excelStartingInteger, 1, apName)  #writes in excel **format** ->(row, column, what is written in cell)
        pyautogui.click(); pyautogui.typewrite(str(formatNumber)) #clicks mouse then types ap number
        pyautogui.press('esc'); pyautogui.press('esc') #hits escape twice
        apNumber += 1 #increments ap number by 1
        excelStartingInteger += 1
        formatNumber = format(apNumber, '05')
    print ('Start of Visio tool') #Uses keypress to fill in names on Visio
#----------------------------------------------------------------------Makes GUI work
    displayText.clear()
    window2.focus()
    displayText.append('Start of Visio tool')
    window3 = Window(app, visible=False)
#----------------------------------------------------------------------Makes GUI work
    while True:
        keyPress = keyboard.read_key() #reads ALL keypress's and saves to variable
        if keyPress == '`' :
            apName = str(apPrefix) + str(formatNumber)
            visioGuts ()
        elif keyPress == 'g':
            apName = str(apPrefix) + str(formatNumber) + 'G'
            visioGuts ()
        elif keyPress == 'h':
            apName = str(apPrefix) + str(formatNumber) + 'H'
            visioGuts ()
        elif keyPress == 'm':
            apName = str(apPrefix) + str(formatNumber) + 'M'
            visioGuts ()
        elif keyPress == 'd':
            apName = str(apPrefix) + str(formatNumber) + 'D'
            visioGuts ()
        elif keyPress == 's':
            apName = str(apPrefix) + str(formatNumber) + 'S'
            visioGuts ()
        elif keyPress == '[': #goes down one ap number.
            apNumber -= 1
            excelStartingInteger -= 1
            formatNumber = format(apNumber, '05')
            apName = str(apPrefix) + str(formatNumber)
            print (apName + ' replace?')
#----------------------------------------------------------------------Makes GUI work
            displayText.clear()
            window2.focus()
            displayText.append(apName + ' replace?')
            window3 = Window(app, visible=False)
#----------------------------------------------------------------------Makes GUI work=
            time.sleep(.08)
        elif keyPress == ']': #goes up one ap number.
            apNumber += 1
            excelStartingInteger += 1
            formatNumber = format(apNumber, '05')
            apName = str(apPrefix) + str(formatNumber)
            print (apName + ' replace?')
#----------------------------------------------------------------------Makes GUI work
            displayText.clear()
            window2.focus()
            displayText.append(apName + ' replace?')
            window3 = Window(app, visible=False)
#----------------------------------------------------------------------Makes GUI work
            time.sleep(.08)
        elif keyPress == 'pause':
            while True:
#----------------------------------------------------------------------Makes GUI work
                displayText.clear()
                window2.focus()
                displayText.append('paused')
                window3 = Window(app, visible=False)
#----------------------------------------------------------------------Makes GUI work
                print('paused')
                time.sleep(.1)
                keyPress = keyboard.read_key() #reads ALL keypress's and saves to variable
                if keyPress == 'pause' :
                    break
#----------------------------------------------------------------------Makes GUI work
            displayText.clear()
            window2.focus()
            displayText.append('unpaused')
            window3 = Window(app, visible=False)
            time.sleep(.1)
            #break
#----------------------------------------------------------------------Makes GUI work
            print('unpaused')
        elif keyPress == '=':
            break
    time.sleep(.1) #prevents cadTool from also breaking
def saveExcel():
    global excelName
    wb.save(excelName) #Saves workbook
    print ('')
    print ('End of Visio Tool')
    print ('')
#----------------------------------------------------------------------Makes GUI work
    displayText.clear()
    window2.focus()
    displayText.append('End of Visio Tool')
    window3 = Window(app, visible=False)
#----------------------------------------------------------------------Makes GUI work
def openExcel(): #Opens excel and reads items to a list
    global number
    global rowNumber
    global name
    global list
    book = openpyxl.load_workbook(excelName)#Opens Excel and declares variables
    sheet = book.active
    while number < 7500: #Reading the excel names and puts them into a list
        rowNumber = str(name) + str(number)
        a1 = sheet[rowNumber]
        list.append(a1)
        number += 1
def cadTool():
#----------------------------------------------------------------------Makes GUI work
    global app
    global window2
    global displayText
    displayText = Text(window2, text=apName, align="left", grid=[0,0])
#----------------------------------------------------------------------Makes GUI work
    print ('Start of CAD tool')
#----------------------------------------------------------------------Makes GUI work
    displayText.clear()
    window2.focus()
    displayText.append('Start of CAD tool')
    window3 = Window(app, visible=False)
#----------------------------------------------------------------------Makes GUI work
    while True: #Start of CAD Tool (reads excel names and places names into CAD)
        global count
        global excelNumber
        test = keyboard.read_key()
        if test == '`' :
            excelNumber = (list[count])
            x = excelNumber.value
            a,b,c = x.split('-')
            newExcelNumber = (a + '-' + b + '\n-' + c)
            print (newExcelNumber)
#----------------------------------------------------------------------Makes GUI work
            displayText.clear()
            window2.focus()
            displayText.append(excelNumber.value)
            window3 = Window(app, visible=False)
#----------------------------------------------------------------------Makes GUI work
            pyautogui.hotkey('ctrl', 'a'); pyautogui.typewrite(newExcelNumber); pyautogui.hotkey('ctrl', 'enter')
            count += 1
        elif test == ']':
            count += 1
            excelNumber = (list[count])
            print (excelNumber.value + ' replace?')
#----------------------------------------------------------------------Makes GUI work
            displayText.clear()
            window2.focus()
            displayText.append(excelNumber.value + ' replace?')
            window3 = Window(app, visible=False)
#----------------------------------------------------------------------Makes GUI work
            time.sleep(.08)
        elif test == '[':
            count -= 1
            excelNumber = (list[count])
            print (excelNumber.value + ' replace?')
#----------------------------------------------------------------------Makes GUI work
            displayText.clear()
            window2.focus()
            displayText.append(excelNumber.value + ' replace?')
            window3 = Window(app, visible=False)
#----------------------------------------------------------------------Makes GUI work
            time.sleep(.08)
        elif test == '=':
            break
    print ('')
    print ('End of CAD Tool')
    print ('')
#----------------------------------------------------------------------Makes GUI work
    displayText.clear()
    window2.focus()
    displayText.append('End of CAD Tool')
    window3 = Window(app, visible=False)
#----------------------------------------------------------------------Makes GUI work

def theGUI():
    def apStartingNumber():
        global apNumber
        apNumber = int(startingNumber.value)
        global formatNumber
        formatNumber = format(apNumber, '05')
        print ('AP number changed to ' + str(apNumber))
        changingText.value = "AP number changed to " + str(apNumber) #Text for changing AP number
    def changeSiteName():
        global siteName
        siteName = str(siteNames.value)
        global apPrefix
        apPrefix = siteName + '-XX-'
        global apName
        apName = str(apPrefix) + str(apNumber)
        print('Site name changed to ' + siteName)
        changingText.value = "Site changed to " + siteName #Text for changing site name
    def visio():
        app.hide()
        window.hide()
        window2.show()
        startWorkbook()
        visioTool()
        saveExcel()
        app.show()
    def cad():
        app.hide()
        window2.show()
        openExcel()
        cadTool()
        app.show()
    def openFileCommand():
        openFiles()
    def open_window():
        window.show()
    def close_window():
        window.hide()

#Main app start
    global app
    app = App(title = "Phoenix_Oath", width=160, height=43, layout='grid')

    button2 = PushButton(app, text = "Visio Tool", command = open_window, align="left", grid=[0,1])
    button3 = PushButton(app, text = "CAD Tool", command = cad, align="right", grid=[2,1])

#First Window start
    window = Window(app, title = "World Destruction", width=352, height=132, layout='grid')
    window.hide()
    button7 = PushButton(window, text = "Go?", command = visio, grid=[2,3])
    
    #Logic for changing starting AP number
    startingNumberText = Text(window, text="AP Number?", align="left", grid=[0,1]) #Text asks for AP number
    startingNumber = TextBox(window, align="right",text = "1", width=30, grid=[1,1]) #Text box for data entry
    button4 = PushButton(window, text = "Confirm", command = apStartingNumber, grid=[2,1])

    #Logic for changing site name
    siteNamesText = Text(window, text="Site Name?", align="left",  grid=[0,2]) #Text asks for site name
    siteNames = TextBox(window, align="right",text = "PAUL", width=30, grid=[1,2]) #Text box for data entry
    button6 = PushButton(window, text = "Confirm", command = changeSiteName, grid=[2,2])

    changingText = Text(window,text="War has changed", align ="left", grid=[1,3])

#Second Window Start
    global window2
    global displayText
    global apName
    window2 = Window(app, title = "World Takeover", width=200, height=45)
    window2.hide()

#Third Window Start
    window3 = Window(app, title = "One bullet at a time", width=200, height=75)
    displayText = Text(window3, text="CAD tool is going", align="left", grid=[0,0]) 
    window3.hide()
    
    openFiles()
    app.display() 

theGUI()




