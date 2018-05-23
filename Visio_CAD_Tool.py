import pyautogui, time, openpyxl, keyboard
from openpyxl import Workbook
from tkinter import filedialog, Tk
from guizero import App, Text, TextBox, PushButton, Window

#Declaring of variables
global siteName
global apPrefix
global apNumber
global apName
global excelCount
global excelNumber
global rowLetter
global rowNumber
global rowWholeName
global list
global excelName
global formatNumber
global excelStartingInteger
global keyPress
global keyPressS
global keyPressG
global keyPressM
global keyPressH
global keyPressD
global keyPressBack
global keyPressForward
global keyPressPause
global keyPressEquals
siteName = str ('Paul')
apPrefix = siteName + '-XX-'
apNumber =  int('1')
apName = str(apPrefix) + str(apNumber)
excelCount = int('1')
excelNumber = int('1')
rowLetter = 'A'
rowNumber = int('1')
rowWholeName = str(rowLetter) + str(rowNumber)
list=[99999]
excelName = 'Paul'
formatNumber = format(apNumber, '05')
excelStartingInteger = int('1')
keyPress = False
keyPressS = False
keyPressG = False
keyPressM = False
keyPressH = False
keyPressD = False
keyPressBack = False
keyPressForward = False
keyPressPause = False
keyPressEquals = False

#Starts an excel. Required for visioTool.
def startWorkbook (): 
    global wb
    wb = Workbook()
    global ws
    ws = wb.active

#Opens file explorer
def openFiles(): 
    global excelName
    Tk().withdraw() # We don't want a full GUI, so keep the root window from appearing
    excelName = filedialog.askopenfilename() # Show an "Open" dialog box and return the path to the selected file
    print(excelName)
    print('')

# This whole function is for reading keypresses. keyboard.read() works better but breaks GUI
# everytime pressed() is called it updates whether a key has been pressed
def pressed(): 
    global keyPress
    global keyPressS
    global keyPressG
    global keyPressM
    global keyPressH
    global keyPressD
    global keyPressBack
    global keyPressForward
    global keyPressPause
    global keyPressEquals
    keyPress = keyboard.is_pressed('`')
    keyPressS = keyboard.is_pressed('s')
    keyPressG = keyboard.is_pressed('g')
    keyPressM = keyboard.is_pressed('m')
    keyPressH = keyboard.is_pressed('h')
    keyPressD = keyboard.is_pressed('d')
    keyPressBack = keyboard.is_pressed('[')
    keyPressForward = keyboard.is_pressed(']')
    keyPressPause = keyboard.is_pressed('pause')
    keyPressEquals = keyboard.is_pressed('=')

# visioTool encapsulates all things Visio Tool related for better organization.
def visioTool():
    def pause():
        global keyPressPause
        keyPressPause = False
        pressed()
        if keyPressPause == True:
            print('Unpaused')
            displayText.clear()
            displayText.append('Unpaused')
            app.after(100,visioLoop)
        else:
            app.after(100,pause)

    def visioGuts():
        global apNumber
        global formatNumber
        global excelStartingInteger
        print (apName)
        ws.cell(excelStartingInteger, 1, apName)  #writes in excel **format** ->(row, column, content to be written in cell)
        pyautogui.press('backspace'); pyautogui.typewrite(str(formatNumber)) #Types ap number
        apNumber += 1 #increments ap number by 1
        excelStartingInteger += 1 # Increments Excel cell to be written in
        formatNumber = format(apNumber, '05') # Modifies apNumber by adding up to 5 zeros in front
        displayText.clear() # Clears GUI
        displayText.append(apName) # Write AP name in GUI
        app.after(1,visioLoop) # after 1ms start the visioLoop
        
    def visioLoop():
        global keyPress
        global keyPressS
        global keyPressG
        global keyPressM
        global keyPressH
        global keyPressD
        global keyPressBack
        global keyPressForward
        global keyPressPause
        global keyPressEquals
        global apNumber
        global formatNumber
        global excelStartingInteger
        global apName
        
        pressed ()
        
        if keyPress == True :
            keyPress = False
            apName = str(apPrefix) + str(formatNumber)
            visioGuts ()
            
        elif keyPressS == True :
            keyPressS = False
            apName = str(apPrefix) + str(formatNumber) + 'S'
            visioGuts ()
            
        elif keyPressG == True :
            keyPressG = False
            apName = str(apPrefix) + str(formatNumber) + 'G'
            visioGuts ()
            
        elif keyPressM == True :
            keyPressM = False
            apName = str(apPrefix) + str(formatNumber) + 'M'
            visioGuts ()
            
        elif keyPressH == True :
            keyPressH = False
            apName = str(apPrefix) + str(formatNumber) + 'H'
            visioGuts ()
            
        elif keyPressD == True :
            keyPressD = False
            apName = str(apPrefix) + str(formatNumber) + 'D'
            visioGuts ()
            
        elif keyPressBack == True: #goes down one ap number.
            keyPressBack = False
            apNumber -= 1
            excelStartingInteger -= 1
            formatNumber = format(apNumber, '05')
            apName = str(apPrefix) + str(formatNumber)
            print (apName)
            displayText.clear()
            displayText.append(apName)
            app.after(80,visioLoop) # 80ms pause before starting the visioLoop to prevent duplication of command
            
        elif keyPressForward == True: #goes up one ap number.
            keyPressForward = False
            apNumber += 1
            excelStartingInteger += 1
            formatNumber = format(apNumber, '05')
            apName = str(apPrefix) + str(formatNumber)
            print (apName)
            displayText.clear()
            displayText.append(apName)
            app.after(80,visioLoop)
            
        elif keyPressPause == True:
            print ('Paused')
            displayText.clear()
            displayText.append('Paused')
            app.after(100,pause)
            
        elif keyPressEquals == True:
            keyPressEquals = False
            saveExcel()
            app.show()
            
        else:
            app.after(1,visioLoop)
            
    global apName
    global apNumber
    global formatNumber
    global excelStartingInteger
    global displayText
    print('Start of Visio Tool')
    displayText.clear()
    displayText.append('Start of Visio Tool')
    visioLoop()

def saveExcel():
    global displayText
    global excelName
    wb.save(excelName) #Saves workbook
    print ('End of Visio Tool')
    displayText.clear()
    displayText.append('End of Visio Tool')
    
def openExcel(): #Opens excel and reads items to a list
    global rowNumber
    global rowWholeName
    global rowLetter
    global list
    book = openpyxl.load_workbook(excelName)#Opens Excel and declares variables
    sheet = book.active #Makes the current sheet active
    for x in range (0,99999): #Reading the excel names and puts them into a list
        rowWholeName = str(rowLetter) + str(rowNumber)
        specificCellValue = sheet[rowWholeName]
        list.append(specificCellValue)
        rowNumber += 1

def cadCapsule():
    def cadTool(): #(reads excel names and places names into CAD)
        global excelCount
        global displayText
        global window2
        global excelNumber
        global keyPress
        global keyPressForward
        global keyPressBack
        global keyPressEquals
        pressed ()
        
        if keyPress == True:
            keyPress = False
            excelNumber = (list[excelCount])
            x = excelNumber.value
            a,b,c = x.split('-') # Splits the ap name into 3 parts
            newExcelNumber = (a + '-' + b + '\n-' + c) # The 3 parts are concatenated together with a character return
            print (newExcelNumber)
            pyautogui.hotkey('ctrl', 'a'); pyautogui.typewrite(newExcelNumber); pyautogui.hotkey('ctrl', 'enter')
            excelCount += 1
            displayText.clear() # Clears the gui
            displayText.append(excelNumber.value) # Displays the AP name
            app.after(1,cadTool)
            
        elif keyPressForward == True: # Looks if the ']' is pressed 
            keyPressForward = False
            excelCount += 1
            excelNumber = (list[excelCount])
            print (excelNumber.value)
            displayText.clear()
            displayText.append(excelNumber.value)
            app.after(80,cadTool)
            
        elif keyPressBack == True:
            keyPressBack = False
            excelCount -= 1
            excelNumber = (list[excelCount])
            print (excelNumber.value)
            displayText.clear()
            displayText.append(excelNumber.value)
            app.after(80,cadTool)
            
        elif keyPressEquals == True:
            keyPressEquals = False
            app.show()
            displayText.clear()
            displayText.append('End of Visio Tool')
            
        else:
            app.after(1,cadTool)
            
    global excelCount
    global displayText
    global window2
    global excelNumber
    global keyPress
    global keyPressForward
    global keyPressBack
    global keyPressEquals
    displayText.clear()
    displayText.append('Start of CAD Tool')
    cadTool()
    
def theGUI():
    def apStartingNumber():
        global apNumber
        global formatNumber
        apNumber = int(startingNumber.value)
        formatNumber = format(apNumber, '05')
        print ('AP number changed to ' + str(apNumber))
        changingText.value = "AP number changed to " + str(apNumber) #Text for changing AP number
    def changeSiteName():
        global siteName
        global apPrefix
        global apName
        siteName = str(siteNames.value)
        apPrefix = siteName + '-XX-'
        apName = str(apPrefix) + str(apNumber)
        print('Site name changed to ' + siteName)
        changingText.value = "Site changed to " + siteName #Text for changing site name
    def visio():
        app.hide()
        window.hide()
        window2.show()
        startWorkbook()
        visioTool()
        #saveExcel()
        #app.show()
    def cad():
        app.hide()
        window2.show()
        openExcel()
        cadCapsule()
        #app.show()
    def openFileCommand():
        openFiles()
    def open_window():
        window.show()
    def close_window():
        window.hide()

#Main app start
    global app
    global window2
    global displayText
    app = App(title = "Phoenix_Oath", width=160, height=43, layout='grid')

    button2 = PushButton(app, text = "Visio Tool", command = open_window, align="left", grid=[0,1])
    button3 = PushButton(app, text = "CAD Tool", command = cad, align="right", grid=[2,1])

#First Window start
    window = Window(app, title = "World Destruction", width=352, height=132, layout='grid')
    window.hide()
    button7 = PushButton(window, text = "Go?", command = visio, grid=[2,3])
    
    #Logic for changing starting AP number
    startingNumberText = Text(window, text="AP Number?", align="left", grid=[0,1]) #Text asks for AP number
    startingNumber = TextBox(window, align="right",text = "1", width=30, grid=[1,1]) #Text box for data entry
    button4 = PushButton(window, text = "Confirm", command = apStartingNumber, grid=[2,1])

    #Logic for changing site name
    siteNamesText = Text(window, text="Site Name?", align="left",  grid=[0,2]) #Text asks for site name
    siteNames = TextBox(window, align="right",text = "PAUL", width=30, grid=[1,2]) #Text box for data entry
    button6 = PushButton(window, text = "Confirm", command = changeSiteName, grid=[2,2])

    changingText = Text(window,text="War has changed", align ="left", grid=[1,3])

#Second Window Start
    window2 = Window(app, title = "World Takeover", width=200, height=25)
    window2.hide()
    displayText = Text(window2, text='Start of Visio Tool', align="left", grid=[0,0])
    
    openFiles()
    app.display()
theGUI()


